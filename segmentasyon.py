#!/usr/bin/env python3
"""
IST Micro - Müşteri Segmentasyon Scripti
Excel'deki 165 kontağı ürün talebine göre 6 kategoriye ayırır.
Çıktı: segmentasyon_raporu.json + api_import_ready.json
"""

import json
import openpyxl
from pathlib import Path

EXCEL_PATH = Path(__file__).parent / "ist_micro_contacts_firma_talep.xlsx"
OUT_PATH = Path(__file__).parent / "segmentasyon_raporu.json"

URUN_KATEGORILERI = {
    "pt100": "Pt100",
    "pt500": "Pt500",
    "pt200": "Pt200",
    "pt1000": "Pt1000",
    "rtd": "Pt Sicaklik Sensorleri",
    "platin": "Pt Sicaklik Sensorleri",
    "sicaklik": "Pt Sicaklik Sensorleri",
    "temperature": "Pt Sicaklik Sensorleri",
    "termokupl": "Termokupl",
    "thermocouple": "Termokupl",
    "thermowell": "Termokupl",
    "ntc": "Termistorler",
    "ptc": "Termistorler",
    "termistor": "Termistorler",
    "basinc": "Basinc Sensorleri",
    "pressure": "Basinc Sensorleri",
    "gaz": "Gaz Akis Sensorleri",
    "akis": "Gaz Akis Sensorleri",
    "flow": "Gaz Akis Sensorleri",
    "mikro isitici": "Mikro Isiticilar",
    "micro heater": "Mikro Isiticilar",
    "heater": "Mikro Isiticilar",
}

def segment_et(ne_istiyor: str, kategori: str, notlar: str) -> list:
    """Bir kontaktan hangi ürün kategorileriyle ilgilendiğini belirle."""
    text = f"{kategori} {ne_istiyor} {notlar}".lower()
    eslesen = set()
    for anahtar, urun_adi in URUN_KATEGORILERI.items():
        if anahtar in text:
            eslesen.add(urun_adi)
    if not eslesen:
        eslesen.add("Genel")
    return sorted(eslesen)


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Tum Kisiler"]

    # Find header row (row 3 has column names)
    headers = {}
    for cell in ws[3]:
        if cell.value:
            col_idx = cell.column - 1  # 0-based
            headers[cell.value] = col_idx

    def val(row, name):
        idx = headers.get(name)
        if idx is None:
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    contacts = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:  # skip empty rows
            continue
        firma = val(row, "Sirket")
        kisi = val(row, "Ad Soyad")
        email = val(row, "E-posta")
        kategori = val(row, "Firma Tipi")
        talep = val(row, "Talep")
        notlar = val(row, "Not")
        unvan = val(row, "Ünvan")
        web = val(row, "Adres / Web")
        ulke = val(row, "Ulke")

        segments = segment_et(talep, kategori, notlar)
        contacts.append({
            "firma": firma,
            "kisi": kisi,
            "email": email,
            "unvan": unvan,
            "web": web,
            "ulke": ulke,
            "kategori": kategori,
            "talep": talep,
            "notlar": notlar,
            "segmentler": segments,
        })

    # İstatistik
    segment_sayilari = {}
    for c in contacts:
        for s in c["segmentler"]:
            segment_sayilari[s] = segment_sayilari.get(s, 0) + 1

    email_var = [c for c in contacts if c["email"]]
    email_yok = [c for c in contacts if not c["email"]]

    rapor = {
        "toplam_kontak": len(contacts),
        "email_olan": len(email_var),
        "email_olmayan": len(email_yok),
        "segment_dagilimi": dict(sorted(segment_sayilari.items(), key=lambda x: -x[1])),
        "kontaklar": contacts,
    }

    OUT_PATH.write_text(json.dumps(rapor, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"✔ Rapor kaydedildi: {OUT_PATH}")
    print(f"  Toplam kontak: {len(contacts)}")
    print(f"  Email olan: {len(email_var)}")
    print(f"  Email olmayan: {len(email_yok)}")
    print(f"  Segmentler:")
    for s, n in sorted(segment_sayilari.items(), key=lambda x: -x[1]):
        print(f"    {s}: {n}")

    # API import için hazır JSON (Customers API'sine POST yapılacak)
    api_import = []
    for c in contacts:
        if c["email"]:
            api_import.append({
                "company": c["firma"],
                "contactPerson": c["kisi"],
                "email": c["email"],
                "country": c["ulke"],
                "category": c["kategori"],
                "request": c["talep"],
                "notes": c["notlar"],
                "tags": ",".join(c["segmentler"]),
                "source": "Sensor+Test 2026 Fuari",
            })
    api_path = Path(__file__).parent / "api_import_ready.json"
    api_path.write_text(json.dumps(api_import, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"\n✔ API import dosyası: {api_path} ({len(api_import)} kayıt)")


if __name__ == "__main__":
    main()
